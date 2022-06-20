# encoding=utf-8
#!/usr/bin/env

import requests
import json
import time
import openpyxl


def SearchAppId(app):
    url = "http://itunes.apple.com/search?term=" + app + "&entity=software"
    r = requests.get(url)
    html = r.content
    html_doc = str(html, 'utf-8')
    data = json.loads(html_doc)
    resultCount = data['resultCount']
    results = data['results']
    print(app + " Find " + str(resultCount) + " result(s)")
    for i in range(resultCount):
        name = results[i]['trackName']
        app_id = results[i]['trackId']
        print("name：" + name, "id：" + str(app_id))


def SaveContent(id, wb, ws):
    row = 2

    for j in range(1, 11):  # 只能爬取前十页
        url = "https://itunes.apple.com/rss/customerreviews/page=" + \
            str(j) + "/id=" + str(id) + "/sortby=mostrecent/json?l=en&&cc=cn"
        r = requests.get(url)

        if r.status_code == 200:
            html = r.content
            html_doc = str(html, 'utf-8')
            data = json.loads(html_doc)["feed"]["entry"]
            for i in data:
                name = i['author']['name']['label']
                rate = i['im:rating']['label']
                user_id = i['id']['label']
                content = i['content']['label']
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=rate)
                ws.cell(row=row, column=3, value=user_id)
                ws.cell(row=row, column=4, value=content)
                row = row + 1
                print(name, rate, user_id, content)
        else:
            return
        # 每一页爬取延迟2秒，以防过于频繁
        time.sleep(2)

def main(appName,appid):
        app = appName
        id = appid.replace('id', '')
        print('====', app, id)
        SearchAppId(app)

        # Workbook init
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="昵称")
        ws.cell(row=1, column=2, value="评分")
        ws.cell(row=1, column=3, value="用户id")
        ws.cell(row=1, column=4, value="评论")

        SaveContent(id, wb, ws)

        wb.save(app + ".xlsx")  # 默认保存在当前目录
        print("Done!")


appName=''
appid=''
# https://apps.apple.com/us/app/indycar/id606905722
#     https://apps.apple.com/us/app/capcut-video-editor/id1500855883
# https://apps.apple.com/cn/app/妙健康-健康管理平台/id841386224?l=ru&see-all=reviews
# https://apps.apple.com/cn/app/%E5%A6%99%E5%81%A5%E5%BA%B7-%E5%81%A5%E5%BA%B7%E7%AE%A1%E7%90%86%E5%B9%B3%E5%8F%B0/id841386224?l=ru&see-all=reviews
apple_app_package_url = os.getenv('apple_app_package_url').strip()
if 'https://apps.apple.com' in apple_app_package_url:
    if '?' in apple_app_package_url:
        apple_app_package_url = apple_app_package_url.split('?')[0]

    appName = apple_app_package_url.split('/')[-2]
    appid = apple_app_package_url.split('/')[-1]
    if not len(appName) > 0:
        print('not support package,', apple_app_package_url, appName)
    print('====', appName, appid)
    main(appName,appid)
